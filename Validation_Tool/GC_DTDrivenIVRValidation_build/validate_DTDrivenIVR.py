
import argparse
import PureCloudPlatformClientV2
from PureCloudPlatformClientV2.rest import ApiException
import os
import openpyxl
from openpyxl.styles import  PatternFill
import configparser
import requests

import pandas as pd

# Create the parser
parser = argparse.ArgumentParser(add_help=False)
# Add an argument
parser.add_argument('--configPath', type=str, required=True, help="Path to <DTDrivenIVR.config>")
# Parse the argument
parser.add_argument('-h', '--help', action='help', default=argparse.SUPPRESS, help="To execute this exe file : .\validate_DTDrivenIVR.exe --configPath <DTDrivenIVR.config> ")
args = parser.parse_args()



#Define columns for objects

path_signal=False
region_signal=False
clientid_signal=False
clientsecret_signal=False
parser = configparser.ConfigParser()
print("Calling config : "+args.configPath)
parser.read(args.configPath)

print ("Initializing...")


if parser.get('Environment', 'region'):
    S_region = PureCloudPlatformClientV2.PureCloudRegionHosts[parser.get('Environment', 'region')]
    region_signal=True
else :
    S_region=''

if parser.get('Environment', 'clientsecret'):
    S_CLIENTSECRET = parser.get('Environment', 'clientsecret')
    clientsecret_signal=True
else :
    S_CLIENTSECRET=''

if parser.get('Environment', 'clientid'):
    S_CLIENTID=parser.get('Environment', 'clientid')
    clientid_signal=True
else :
    S_CLIENTID=''


if parser.get('Environment', 'csv'):
    S_DT_CSV=parser.get('Environment', 'csv')
    path_signal=True
else :
    S_DT_CSV=''


if parser.get('Objects', 'skill_col_list'):
    skill_col_list=parser.get('Objects', 'skill_col_list').split(',')
else :
    skill_col_list=[]
if parser.get('Objects', 'queue_col_list'):
    queue_col_list=parser.get('Objects', 'queue_col_list').split(',')
else :
    queue_col_list=[]
if parser.get('Objects', 'prompt_col_list'):
    prompt_col_list=parser.get('Objects', 'prompt_col_list').split(',')
else :
    prompt_col_list=[]
if parser.get('Objects', 'schedulegroup_col_list'):
    schedulegroup_col_list=parser.get('Objects', 'schedulegroup_col_list').split(',')
else :
    schedulegroup_col_list=[]
if parser.get('Objects', 'schedule_col_list'):
    schedule_col_list=parser.get('Objects', 'schedule_col_list').split(',')
else :
    schedule_col_list=[]
if parser.get('Objects', 'emergencygroup_col_list'):
    emergencygroup_col_list=parser.get('Objects', 'emergencygroup_col_list').split(',')
else :
    emergencygroup_col_list=[]

if os.path.exists(S_DT_CSV):
    path_signal=True


#Store Compare Objects/Error for summmary
store_key_error=[]
store_skill_error=[]
store_queue_error=[]
store_prompt_error=[]
store_schedulegroup_error=[]
store_schedule_error=[]
store_emergencygroup_error=[]

#Store API objects
skill_data=[]
prompt_data=[]
queue_data=[]
schedule_data=[]
schedulegroup_data=[]
emergencygroup_data=[]
catalogue={
            "skill":{
                "name":'ACDSkill',
                "sheet_name":'ACDSkill',
                "color":'5cb800',
                "data":'skill_data' #API skill reference
            },
            "queue":{
                "name":'Queue',
                "sheet_name":'Queue',
                "color":'ff3333',
                "data":'queue_data' #API queue reference
            },
            "prompt":{
                "name":'Prompt',
                "sheet_name":'Prompt',
                "color":'ff4dc4',
                "data":'prompt_data' #API queue reference
            },
            "schedulegroup":{
                "name":'ScheduleGroup',
                "sheet_name":'ScheduleGroup',
                "color":'6666ff',
                "data":'schedulegroup_data' #API queue reference
            },
            "schedule":{
                "name":'Schedule',
                "sheet_name":'Schedule',
                "color":'cccc00',
                "data":'schedule_data' #API queue reference
            },
            "emergencygroup":{
                "name":'EmergencyGroup',
                "sheet_name":'EmergencyGroup',
                "color":'ff9900',
                "data":'schedule_data' #API queue reference
            },
        }



#if args.region:
if (S_region and S_CLIENTID and S_CLIENTSECRET and path_signal):
    print("Setting up Environment Variables : ")
    print("ClientID : "+S_CLIENTID)
    print("ClientSecret : "+S_CLIENTSECRET)
    print("Region : "+S_region.name)
    print("CSV : "+S_DT_CSV)
    

    print("===================START======================")
     
    def CheckORCreateDir(OUTPUT_PATH):
        print("Checking OUTPUT dir :" + OUTPUT_PATH)
        MYDIR = (OUTPUT_PATH)
        CHECK_FOLDER = os.path.isdir(MYDIR)
        if not CHECK_FOLDER:
            os.makedirs(MYDIR)
            print("Created folder : ", MYDIR)
        else:
            print(MYDIR, "Folder already exists.")

    
    def delfile(file) :
        #print("Checking old file : "+ file)
        if os.path.exists(file):
            os.remove(file)
            print("Found old file and removed : "+ file)
    

    def downloadWAV(filename,link):
        print("Downloading audio ")
        x = requests.get(link)

        with open(filename, 'wb') as f:
            f.write(x.content)
            f.close
        print("Done downloading audio file : "+filename)



    def GC_apiclient(clientId,clientSecret,region):
        PureCloudPlatformClientV2.configuration.host = region.get_api_host()
        apiclient = PureCloudPlatformClientV2.api_client.ApiClient().get_client_credentials_token(clientId,clientSecret)
        return apiclient

    def getPrompts_list(page_number):
        page_size = 100 # int | Page size (optional) (default to 25)
        sort_by = 'id' # str | Sort by (optional) (default to 'id')
        sort_order = 'asc' # str | Sort order (optional) (default to 'asc')
        try:
            # Get a pageable list of flows, filtered by query parameters
            api_response = architectApi.get_architect_prompts(
                                                                page_number=page_number, 
                                                                page_size=page_size, 
                                                                sort_by=sort_by, 
                                                                sort_order=sort_order
                                                                )
            #print(api_response)
            return api_response
        except ApiException as e:
            print("Exception when calling GetFlowsRequest->getPrompts_list: %s\n" % e)


    def getACDSKills_list(page_number):
        page_size = 100 # int | Page size (optional) (default to 25)
        try:
            # Get a pageable list of flows, filtered by query parameters
            api_response = routingApi.get_routing_skills(
                                                            page_number=page_number, 
                                                            page_size=page_size
                                                        )
            #print(api_response)
            return api_response
        except ApiException as e:
            print("Exception when calling GetFlowsRequest->getACDSKills_list: %s\n" % e)


    def getACDSQueues_list(page_number):
        page_size = 100 # int | Page size (optional) (default to 25)
        try:
            # Get a pageable list of flows, filtered by query parameters
            api_response = routingApi.get_routing_queues(
                                                            page_number=page_number, 
                                                            page_size=page_size
                                                        )
            #print(api_response)
            return api_response
        except ApiException as e:
            print("Exception when calling GetFlowsRequest->getACDSQueues_list: %s\n" % e)

    def getScheduleGroups_list(page_number):
        page_size = 100 # int | Page size (optional) (default to 25)
        try:
            # Get a pageable list of flows, filtered by query parameters
            api_response = architectApi.get_architect_schedulegroups(
                                                            page_number=page_number, 
                                                            page_size=page_size
                                                        )
            #print(api_response)
            return api_response
        except ApiException as e:
            print("Exception when calling GetFlowsRequest->getScheduleGroups_list: %s\n" % e)

    def getSchedules_list(page_number):
        page_size = 100 # int | Page size (optional) (default to 25)
        try:
            # Get a pageable list of flows, filtered by query parameters
            api_response = architectApi.get_architect_schedules(
                                                            page_number=page_number, 
                                                            page_size=page_size
                                                        )
            #print(api_response)
            return api_response
        except ApiException as e:
            print("Exception when calling GetFlowsRequest->getSchedules_list: %s\n" % e)

    def getEmergencyGroups_list(page_number):
        page_size = 100 # int | Page size (optional) (default to 25)
        try:
            # Get a pageable list of flows, filtered by query parameters
            api_response = architectApi.get_architect_emergencygroups(
                                                            page_number=page_number, 
                                                            page_size=page_size
                                                        )
            #print(api_response)
            return api_response
        except ApiException as e:
            print("Exception when calling GetFlowsRequest->getEmergencyGroups_list: %s\n" % e)




    #Convert CSV to XLSX
    def csv2xlsx(filename):
        read_file = pd.read_csv (filename)
        new_filename=filename.replace('.csv','.xlsx')
        read_file.to_excel (new_filename, index = None, header=True)
        return new_filename

    #Duplicate Worksheet
    def create_AnalaysingSheet(filename):
        
        wb = openpyxl.load_workbook (filename)
        print("create ...Analysis Sheet")
        ori_ws = wb['Sheet1']    
        ori_ws.title = 'Original'
        analyzeSheet=wb.copy_worksheet(ori_ws)
        analyzeSheet.title='AnalaysingSheet'
        analyzeSheet.sheet_properties.tabColor = 'FFFF00'
        for cell in analyzeSheet[1:1]:
                    cell.fill = PatternFill("solid", start_color="00b8e6")
        c = analyzeSheet['B2']
        analyzeSheet.freeze_panes = c
        
        print("create ...Summary Sheet as placeholder")
        wb.create_sheet("Summary")
        wb.save(filename)

    #Save Extraction into Worksheet
    def create_ExtractionSheet(catalogue_obj,filename,data):
        wb = openpyxl.load_workbook (filename)
        print("create ..."+catalogue_obj["sheet_name"]+" Sheet")
        sheet = wb.create_sheet(catalogue_obj["sheet_name"])
        sheet.sheet_properties.tabColor=catalogue_obj["color"]
        sheet['A1']=catalogue_obj["name"]
        sheet['A1'].fill = PatternFill("solid", start_color=catalogue_obj["color"])
        count=2
        for row in data:
            join="A"+str(count)
            #print(join)
            sheet[join] = row
            count+=1   
        wb.save(filename)
        print("done")

    def chk_missing_key_AnalaysingSheet(filename,errordata,errorcolor):
    #Change the Sheet name 
        wb = openpyxl.load_workbook (filename)
        print("Check ...missing key in AnalaysingSheet")
        
        ws_AS=wb['AnalaysingSheet']
        chk_key_col='A'
        count=0
        for col in ws_AS[chk_key_col]:
                count+=1
                if(count>1):
                    if col.value:
                        continue 
                        
                    else:
                        errordata.append(str(count))
                        for cell in ws_AS[str(count):str(count)]:
                            #Fill RED color
                            cell.fill = PatternFill("solid", start_color=errorcolor)
        wb.save(filename)
        if(len(errordata)>0):
            print("Found and Highlighted missing key at row : "+str(errordata))
        else : 
            print("No missing key")


    def chk_mismatch_object(filename,catalogue_obj,input_object_col,object_data,errordata,errorcolor):
        wb = openpyxl.load_workbook (filename)
        ws_AS=wb['AnalaysingSheet']
        print("Check ...mismatch "+catalogue_obj["name"]+" in AnalaysingSheet cell(s)")
        print("Pre-set "+catalogue_obj["name"]+" col : "+ str(input_object_col))
        for obj_col in input_object_col:
            count=0
            for col in ws_AS[obj_col]:
                count+=1
                if(count>1):
                    if col.value:
                        Found=False
                        for skill in object_data:
                            if(skill == col.value):
                                Found = True
                                #print(obj_col+str(count)+" "+col.value)
                        if(Found == False):
                            errordata.append(obj_col+str(count)+"||"+col.value)
                            ws_AS[obj_col+str(count)].fill = PatternFill("solid", start_color=errorcolor)
                            #print(obj_col+str(count)+" "+col.value)
                    
        wb.save(filename)
        if(len(errordata)>0):
                print("Found and Highlighted mismatch "+catalogue_obj["name"]+" at cell(s), Total : "+str(len(errordata)))
        else : 
                print("No mismatch "+catalogue_obj["name"]+" found")
        
    def list_object_summary(filename,catalogue_obj,pos_col,name_col,total_col,error_data):

        wb = openpyxl.load_workbook (filename)
        print("open ...Summary Sheet for "+catalogue_obj["name"])
        ws_S=wb['Summary']
        ws_S.sheet_properties.tabColor="66ffff"
        ws_S[pos_col+'1']="Error_"+catalogue_obj["name"]+"_Position"
        ws_S[pos_col+'1'].fill = PatternFill("solid", start_color=catalogue_obj["color"])
                            
        ws_S[name_col+'1']="Error_"+catalogue_obj["name"]+"_Name"
        ws_S[name_col+'1'].fill = PatternFill("solid", start_color=catalogue_obj["color"])
        ws_S[total_col+'1']="TotalErr_"+catalogue_obj["name"]+"="+str(len(error_data))
        ws_S[total_col+'1'].fill = PatternFill("solid", start_color=catalogue_obj["color"])
        count=2
        for row in error_data:
            #print(join)
            val=row.split("||")
            join=pos_col+str(count)
            ws_S[join] =val[0]
            ws_S[join].fill = PatternFill("solid", start_color=catalogue_obj["color"])
            join=name_col+str(count)
            ws_S[join] =val[1]
            ws_S[join].fill = PatternFill("solid", start_color=catalogue_obj["color"])
            count+=1   
        wb.save(filename)
        print("done listing "+catalogue_obj["name"]+" error")
    apiclient=GC_apiclient(S_CLIENTID,S_CLIENTSECRET,S_region)
    architectApi = PureCloudPlatformClientV2.ArchitectApi(apiclient)
    routingApi = PureCloudPlatformClientV2.RoutingApi(apiclient)

    print("==============API Extracting Skills==================")
    print("Python SDK retrieving Org Skills : ")
    data = getACDSKills_list(0)
    print("Number of Pages : "+ str(data.page_count))
    print("Number Skills per page : "+str(data.page_size))
    print("Total Skills : "+str(data.total))
    #print(data)
    print("Storing Skill into skill_data")
    count = 0 
    for y in range(data.page_count):

        data = getACDSKills_list(y+1)
        print("Loading page "+ str(int(y+1)) +" into skill_data...")
        for x in data.entities:
            count = count+1
            #print(str(count) +" Skill :"+ str(x.name))
            name = x.name
            skill_data.append(x.name)
            #print(skill_data)
    print("*Complete* storing Skill")
    
    print("==============API Extracting Queues==================")
    print("Python SDK retrieving Org Queues : ")
    data = getACDSQueues_list(0)
    print("Number of Pages : "+ str(data.page_count))
    print("Number Queues per page : "+str(data.page_size))
    print("Total Queues : "+str(data.total))
    #print(data)
    print("Storing Queue into queue_data")
    count = 0 
    for y in range(data.page_count):
        data = getACDSQueues_list(y+1)
        print("Loading page "+ str(int(y+1)) +" into queue_data...")
        for x in data.entities:
            count = count+1
            #print(str(count) +" Queue :"+ str(x.name))
            name = x.name
            queue_data.append(x.name)
    print("*Complete* storing Queue")        

    print("==============API Extracting Prompts==================")
    print("Python SDK retrieving Org Prompts : ")
    data = getPrompts_list(0)
    print("Number of Pages : "+ str(data.page_count))
    print("Number Prompts per page : "+str(data.page_size))
    print("Total Prompts : "+str(data.total))
    #print(data)
    print("Storing Prompt into prompt_data")
    count = 0 
    for y in range(data.page_count):
        data = getPrompts_list(y+1)
        print("Loading page "+ str(int(y+1)) +" into prompt_data...")
        for x in data.entities:
            count = count+1
            #print(str(count) +" Prompt :"+ str(x.name))
            name = x.name
            prompt_data.append(x.name)
    print("*Complete* storing Prompt")  

    print("==============API Extracting ScheduleGroups==================")
    print("Python SDK retrieving Org ScheduleGroups : ")
    data = getScheduleGroups_list(0)
    print("Number of Pages : "+ str(data.page_count))
    print("Number ScheduleGroups per page : "+str(data.page_size))
    print("Total ScheduleGroups : "+str(data.total))
    #print(data)
    print("Storing ScheduleGroup into schedulegroup_data")
    count = 0 
    for y in range(data.page_count):
        data = getScheduleGroups_list(y+1)
        print("Loading page "+ str(int(y+1)) +" into schedulegroup_data...")
        for x in data.entities:
            count = count+1
            #print(str(count) +" ScheduleGroup :"+ str(x.name))
            name = x.name
            schedulegroup_data.append(x.name)
    print("*Complete* storing ScheduleGroup")        


    print("==============API Extracting Schedules==================")
    print("Python SDK retrieving Org Schedules : ")
    data = getSchedules_list(0)
    print("Number of Pages : "+ str(data.page_count))
    print("Number Schedules per page : "+str(data.page_size))
    print("Total Schedules : "+str(data.total))
    #print(data)
    print("Storing Schedule into schedule_data")
    count = 0 
    for y in range(data.page_count):
        data = getSchedules_list(y+1)
        print("Loading page "+ str(int(y+1)) +" into schedule_data...")
        for x in data.entities:
            count = count+1
            #print(str(count) +" Schedule :"+ str(x.name))
            name = x.name
            schedule_data.append(x.name)
    print("*Complete* storing Schedule")         
    

    print("==============API Extracting EmergencyGroups==================")
    print("Python SDK retrieving Org EmergencyGroups : ")
    data = getEmergencyGroups_list(0)
    print("Number of Pages : "+ str(data.page_count))
    print("Number EmergencyGroups per page : "+str(data.page_size))
    print("Total EmergencyGroups : "+str(data.total))
    #print(data)
    print("Storing Schedule into emergencygroup_data")
    count = 0 
    for y in range(data.page_count):
        data = getEmergencyGroups_list(y+1)
        print("Loading page "+ str(int(y+1)) +" into emergencygroup_data...")
        for x in data.entities:
            count = count+1
            #print(str(count) +" Schedule :"+ str(x.name))
            name = x.name
            emergencygroup_data.append(x.name)
    print("*Complete* storing EmergencyGroup")         
    



    #Copy from .csv to .xlsx
    print("==============Creating Sheets==================")
    newfilename=csv2xlsx(S_DT_CSV)
    create_AnalaysingSheet(newfilename)
    create_ExtractionSheet(catalogue["skill"],newfilename,skill_data)
    create_ExtractionSheet(catalogue["queue"],newfilename,queue_data)
    create_ExtractionSheet(catalogue["prompt"],newfilename,prompt_data)
    create_ExtractionSheet(catalogue["schedulegroup"],newfilename,schedulegroup_data)
    create_ExtractionSheet(catalogue["schedule"],newfilename,schedule_data)
    create_ExtractionSheet(catalogue["emergencygroup"],newfilename,emergencygroup_data)
    
    
    
    
    print("==============Perform Sheet Analysis==================")
    #1st rule : Check missing Key
    chk_missing_key_AnalaysingSheet(newfilename,store_key_error,"ff0000")

    chk_mismatch_object(newfilename,catalogue["skill"],skill_col_list,skill_data,store_skill_error,catalogue["skill"]["color"])
    chk_mismatch_object(newfilename,catalogue["queue"],queue_col_list,queue_data,store_queue_error,catalogue["queue"]["color"])
    chk_mismatch_object(newfilename,catalogue["prompt"],prompt_col_list,prompt_data,store_prompt_error,catalogue["prompt"]["color"])
    chk_mismatch_object(newfilename,catalogue["schedulegroup"],schedulegroup_col_list,schedulegroup_data,store_schedulegroup_error,catalogue["schedulegroup"]["color"])
    chk_mismatch_object(newfilename,catalogue["schedule"],schedule_col_list,schedule_data,store_schedule_error,catalogue["schedule"]["color"])
    chk_mismatch_object(newfilename,catalogue["emergencygroup"],emergencygroup_col_list,emergencygroup_data,store_emergencygroup_error,catalogue["emergencygroup"]["color"])
    
    print("==============Summarizing Sheet Analysis==================")
    list_object_summary(newfilename,catalogue["skill"],'A','B','C',store_skill_error)
    list_object_summary(newfilename,catalogue["queue"],'D','E','F',store_queue_error)
    list_object_summary(newfilename,catalogue["prompt"],'G','H','I',store_prompt_error)
    list_object_summary(newfilename,catalogue["schedulegroup"],'J','K','L',store_schedulegroup_error)
    list_object_summary(newfilename,catalogue["schedule"],'M','N','O',store_schedule_error)
    list_object_summary(newfilename,catalogue["emergencygroup"],'P','Q','R',store_emergencygroup_error)
    

    print("===================END======================")
else:
    print("Invalid configuration in DTDrivenIVR.config")
    if(path_signal==False):
        print("--> Invalid filepath.")
    if(region_signal==False):
        print("--> Invalid region.")
    if(clientid_signal==False):
        print("--> Invalid clientid.")
    if(clientsecret_signal==False):
        print("--> Invalid clientsecret.")